# -*- coding: utf-8 -*-
"""
excel_generator.py

CLI 生成：
- 词汇笔记.xlsx（单 sheet）  <-- 从 global_cache 导出
- shuffle_e2c_master.xlsx（多 sheet，每 sheet N 条） <-- 从“本次词汇笔记 Excel”读入后乱序导出
  同时写出 meta.json，用于前端隐藏无效练习选项（例如不存在例句字段就不显示例句练习）。

兼容 field_definitions.get_*_excel_columns() 两种返回：
- List[FieldDef]  (有 .key + 表头属性可能叫 title/header/name/label/...)
- List[str]       (字段 key 列表，需要从 fd.FIELD_DEFS 取表头)

关键规则（按你的最新决策）：
1) shuffle_e2c 必须以“本次词汇笔记 Excel”为输入源，保证 A=本次输入（不从 global_cache 全量乱序）。
2) shuffle 输出列 = vocab_note present_keys（实际存在字段） + 强制追加：no, word_original, audio_primary, timer, combo
   - timer/combo 为 computed：即使词汇笔记没有，也要写出来
3) 如果词汇笔记缺少 example/example_cn，则 meta.json 不会包含 sent_e2c/sent_c2e，前端应隐藏例句选项。
"""

from __future__ import annotations

import argparse
import json
import os
import random
import sys
from pathlib import Path
from typing import Any, Dict, List, Sequence, Union, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# -----------------------------
# 让 "from core import ..." 在直接运行脚本时也稳定
# -----------------------------
REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from core import field_definitions as fd  # noqa
from core.entry_schema import Entry  # noqa

EntryLike = Union[Entry, Dict[str, Any]]


# -----------------------------
# utils
# -----------------------------
def _ensure_parent_dir(path: str) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)


def _load_json(path: str) -> Dict[str, Any]:
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _flatten_cache_entries(cache_obj: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    兼容缓存结构：
      { word_original: { entries: [ {entry}, ... ] }, ... }
    """
    out: List[Dict[str, Any]] = []
    if not isinstance(cache_obj, dict):
        return out
    for _, group in cache_obj.items():
        if not isinstance(group, dict):
            continue
        entries = group.get("entries", [])
        if isinstance(entries, list):
            for e in entries:
                if isinstance(e, dict):
                    out.append(e)
    return out


def _get_entry_value(entry: EntryLike, key: str) -> Any:
    if isinstance(entry, dict):
        return entry.get(key, "")
    return getattr(entry, key, "")


def _format_cell_value(key: str, value: Any) -> Any:
    if value is None:
        return ""
    if key == "synonyms":
        if isinstance(value, list):
            return "; ".join(str(x).strip() for x in value if str(x).strip())
        return str(value)
    return str(value)


def _write_header(ws: Worksheet, titles: Sequence[str]) -> None:
    ws.append(list(titles))


def _compute_timer_value() -> int:
    fn = getattr(fd, "compute_timer_value", None)
    if callable(fn):
        try:
            return int(fn())
        except Exception:
            return 5000
    return 5000


def _compute_combo_value(word: str) -> Any:
    fn = getattr(fd, "compute_combo_value", None)
    if callable(fn):
        try:
            return fn(word)
        except Exception:
            return 2
    return 2


def _write_rows(ws: Worksheet, entries: Sequence[EntryLike], col_keys: Sequence[str]) -> None:
    valid_computed = getattr(fd, "VALID_COMPUTED_FIELD_SET", set(["timer", "combo"]))
    for e in entries:
        row: List[Any] = []
        for k in col_keys:
            if k in valid_computed:
                if k == "timer":
                    row.append(_compute_timer_value())
                elif k == "combo":
                    word = _get_entry_value(e, "word_original")
                    row.append(_compute_combo_value(str(word)))
                else:
                    row.append("")
            else:
                if k == "word_original":
                    # 导出展示优先用 word_display（导出层生成的字段）
                    v = (
                        _get_entry_value(e, "word_display")
                        or _get_entry_value(e, "word_norm")
                        or _get_entry_value(e, "word_original")
                    )
                else:
                    v = _get_entry_value(e, k)

                row.append(_format_cell_value(k, v))
        ws.append(row)


def _chunk(seq: Sequence[Any], size: int) -> List[Sequence[Any]]:
    if size <= 0:
        return [seq]
    return [seq[i: i + size] for i in range(0, len(seq), size)]


def _resolve_selected_fields(preset: str) -> List[str]:
    fn = getattr(fd, "resolve_preset", None)
    if callable(fn):
        return list(fn(preset))
    # fallback
    return [
        "word_original",
        "phonetic_uk",
        "phonetic_us",
        "pos_cn",
        "definition_en",
        "example",
        "example_cn",
        "synonyms",
    ]


def _pick_title(obj: Any, default: str) -> str:
    """
    兼容你本地 FieldDef 的表头字段名可能不是 title。
    """
    for attr in ("title", "header", "name", "label", "col", "column"):
        if hasattr(obj, attr):
            v = getattr(obj, attr)
            if v is None:
                continue
            s = str(v).strip()
            if s:
                return s
    return default


def _normalize_columns(cols: Any) -> List[Dict[str, str]]:
    """
    统一成：
      [ {"key": "...", "title": "..."}, ... ]
    """
    if not isinstance(cols, list):
        raise TypeError(f"columns must be list, got: {type(cols)}")

    if not cols:
        return []

    first = cols[0]

    # Case A: FieldDef 列表
    if hasattr(first, "key"):
        out: List[Dict[str, str]] = []
        for c in cols:
            k = str(getattr(c, "key")).strip()
            out.append({"key": k, "title": _pick_title(c, k)})
        return out

    # Case B: str 列表
    if isinstance(first, str):
        field_defs = getattr(fd, "FIELD_DEFS", {})
        out2: List[Dict[str, str]] = []
        for k in cols:
            k = str(k).strip()
            obj = field_defs.get(k)
            if obj is None:
                out2.append({"key": k, "title": k})
            else:
                out2.append({"key": k, "title": _pick_title(obj, k)})
        return out2

    raise TypeError(f"Unsupported column item type: {type(first)}")


# -----------------------------
# vocab_note export (from global_cache)
# -----------------------------
def export_vocab_note_excel(
    entries: Sequence[EntryLike],
    selected_fields: List[str],
    output_path: str,
    sheet_name: str = "vocab",
) -> str:
    _ensure_parent_dir(output_path)

    raw_cols = fd.get_vocab_note_excel_columns(selected_fields)
    cols = _normalize_columns(raw_cols)

    col_keys = [c["key"] for c in cols]
    titles = [c["title"] for c in cols]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    _write_header(ws, titles)
    _write_rows(ws, entries, col_keys)

    wb.save(output_path)
    return output_path


# -----------------------------
# read entries from vocab_note excel
# -----------------------------
def _build_title_to_key_map() -> Dict[str, str]:
    """
    建立“表头文本 -> key”的映射，兼容：
    - 表头是 key
    - 表头是 fd.FIELD_DEFS 的 title/header/name/label...
    """
    m: Dict[str, str] = {}
    field_defs = getattr(fd, "FIELD_DEFS", {})

    # key & title 双向兼容
    for k, obj in field_defs.items():
        k2 = str(k).strip()
        if not k2:
            continue
        m[k2.lower()] = k2
        title = _pick_title(obj, k2)
        if title:
            m[str(title).strip().lower()] = k2

    # 常见中文表头兜底（防你改 title 或历史文件表头不同）
    m["英文单词"] = "word_original"
    m["中文解释"] = "pos_cn"
    m["例句"] = "example"
    m["例句翻译"] = "example_cn"
    m["学生答案"] = "student_answer"
    m["音频"] = "audio_primary"
    return m

import re

def _slugify_word_for_audio(word: str) -> str:
    s = (word or "").strip().lower()
    # 空格、斜杠等变下划线
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[\/\\]+", "_", s)
    # 去掉括号等标点，只留字母数字下划线
    s = re.sub(r"[^a-z0-9_]+", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s

def ensure_audio_primary(entry: Dict[str, Any]) -> None:
    """
    保证 entry['audio_primary'] 有值：
    - 优先用原有字段
    - 否则按统一规则生成 /static/audio/uk/{slug}.mp3
    """
    ap = (entry.get("audio_primary") or "").strip()
    if ap:
        return
    w = (entry.get("word_original") or "").strip()
    if not w:
        entry["audio_primary"] = ""
        return
    entry["audio_primary"] = f"/static/audio/uk/{_slugify_word_for_audio(w)}.mp3"


def load_entries_from_vocab_excel(vocab_excel_path: str, sheet_name: str = "vocab") -> Dict[str, Any]:
    """
    从“本次词汇笔记 Excel”读取 entry dict 列表，并返回 present_keys。
    返回：
      { "entries": [...], "present_keys": [...] }
    """
    if not os.path.exists(vocab_excel_path):
        raise FileNotFoundError(vocab_excel_path)

    wb = load_workbook(vocab_excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"entries": [], "present_keys": []}

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    title2key = _build_title_to_key_map()

    col_keys: List[str] = []
    for h in header:
        k = title2key.get(h.strip().lower())
        col_keys.append(k or "")

    present_keys = sorted({k for k in col_keys if k})
    entries: List[Dict[str, Any]] = []

    for r in rows[1:]:
        if r is None:
            continue
        if all((c is None or str(c).strip() == "") for c in r):
            continue

        e: Dict[str, Any] = {}
        for idx, k in enumerate(col_keys):
            if not k:
                continue
            if idx >= len(r):
                continue
            v = r[idx]
            e[k] = "" if v is None else str(v).strip()
        entries.append(e)

    return {"entries": entries, "present_keys": present_keys}


# -----------------------------
# meta.json for front-end options
# -----------------------------
def compute_available_practice_types(present_keys: Sequence[str]) -> List[str]:
    """
    返回可展示的练习类型：
      word_e2c, word_c2e, sent_e2c, sent_c2e
    """
    keys = set(present_keys)
    avail: List[str] = []
    if "word_original" in keys:
        avail.append("word_e2c")
    if "pos_cn" in keys:
        avail.append("word_c2e")
    if "example" in keys:
        avail.append("sent_e2c")
    if "example_cn" in keys:
        avail.append("sent_c2e")
    return avail


def write_meta_json(meta_path: str, present_keys: Sequence[str]) -> str:
    obj = {
        "present_keys": list(present_keys),
        "available_practice_types": compute_available_practice_types(present_keys),
    }
    _ensure_parent_dir(meta_path)
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
    return meta_path


# -----------------------------
# shuffle export (from vocab_note excel)
# -----------------------------
def export_shuffle_e2c_excel_from_present_keys(
    shuffled_entries: Sequence[EntryLike],
    present_keys: Sequence[str],
    output_path: str,
    base_sheet_name: str = "shuffle_e2c",
    max_rows_per_sheet: int = 25,
) -> str:
    """
    shuffle master 输出列：
    - 以 vocab_note 的 present_keys 为主（不凭空生成缺失字段）
    - 但强制补齐练习契约字段：no, word_original, audio_primary, timer, combo
    - **表头一律写 key（全局统一）**
    """
    _ensure_parent_dir(output_path)

    valid_out = getattr(fd, "VALID_OUTPUT_FIELD_SET", None)
    keys = [k for k in present_keys if (valid_out is None or k in valid_out)]

    must = ["no", "word_original", "audio_primary", "timer", "combo"]
    for m in must:
        if m not in keys:
            keys.append(m)

    cols = _normalize_columns(keys)
    col_keys = [c["key"] for c in cols]  # ✅ 只用 key

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    chunks = _chunk(list(shuffled_entries), max_rows_per_sheet)
    for i, part in enumerate(chunks, start=1):
        name = f"{base_sheet_name}_{i}" if len(chunks) > 1 else base_sheet_name
        ws = wb.create_sheet(title=name)

        # ✅ 表头写 key（统一规范）
        _write_header(ws, col_keys)

        _write_rows(ws, part, col_keys)

    wb.save(output_path)
    return output_path

# -----------------------------
# main
# -----------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["vocab_note", "shuffle_e2c"], required=True)

    # vocab_note: still from cache
    parser.add_argument("--global_cache", default="", help="global_cache.json（vocab_note 模式需要）")

    # shuffle_e2c: must from vocab excel
    parser.add_argument("--vocab_excel", default="", help="本次词汇笔记 Excel（shuffle_e2c 模式必须）")

    parser.add_argument("--output", required=True, help="输出 xlsx 路径")
    parser.add_argument("--meta_output", default="", help="输出 meta.json 路径（默认与 output 同名 .meta.json）")

    parser.add_argument("--preset", default="all", help="词汇笔记字段预设（vocab_note 用）")
    parser.add_argument("--sheet_name", default="vocab", help="词汇笔记 sheet 名（读/写）")
    parser.add_argument("--max_rows_per_sheet", type=int, default=25)
    parser.add_argument("--base_sheet_name", default="shuffle_e2c", help="shuffle sheet base name")
    parser.add_argument("--seed", type=int, default=0, help="shuffle 随机种子（0 表示不固定）")
    args = parser.parse_args()

    if args.mode == "vocab_note":
        if not args.global_cache:
            raise SystemExit("[FATAL] vocab_note 模式必须提供 --global_cache")

        cache = _load_json(args.global_cache)
        entries = _flatten_cache_entries(cache)
        if not entries:
            raise SystemExit(f"[ERROR] No entries found in cache: {args.global_cache}")

        selected = _resolve_selected_fields(args.preset)
        out = export_vocab_note_excel(entries, selected, args.output, sheet_name=args.sheet_name)
        print(f"[DONE] vocab_note -> {os.path.abspath(out)}  rows={len(entries)}")
        return

    # shuffle_e2c
    print("[NOTICE] 生产配置建议：不要对用户开放 TEMP_ONLY / FORCE_OVERWRITE；shuffle 将严格以本次 vocab_note.xlsx 为输入源。")

    if not args.vocab_excel:
        raise SystemExit("[FATAL] shuffle_e2c 模式必须提供 --vocab_excel（本次词汇笔记 Excel），以保证 A=本次输入。")

    pack = load_entries_from_vocab_excel(args.vocab_excel, sheet_name=args.sheet_name)
    entries = pack["entries"]
    present_keys = pack["present_keys"]
    if not entries:
        raise SystemExit(f"[ERROR] No entries found in vocab_excel: {args.vocab_excel}")

    # meta.json
    meta_path = args.meta_output.strip() or (os.path.splitext(args.output)[0] + ".meta.json")
    write_meta_json(meta_path, present_keys)
    print(f"[INFO] meta -> {os.path.abspath(meta_path)}  available={compute_available_practice_types(present_keys)}")

    # shuffle
    entries2 = list(entries)
    if args.seed and args.seed != 0:
        rnd = random.Random(args.seed)
        rnd.shuffle(entries2)
    else:
        random.shuffle(entries2)
    # ✅ 新增：补齐 audio_primary
    for e in entries2:
        if isinstance(e, dict):
            ensure_audio_primary(e)

    out = export_shuffle_e2c_excel_from_present_keys(
        shuffled_entries=entries2,
        present_keys=present_keys,
        output_path=args.output,
        base_sheet_name=args.base_sheet_name,
        max_rows_per_sheet=args.max_rows_per_sheet,
    )
    print(f"[DONE] shuffle_e2c(from vocab_excel) -> {os.path.abspath(out)}  rows={len(entries2)}")


if __name__ == "__main__":
    main()