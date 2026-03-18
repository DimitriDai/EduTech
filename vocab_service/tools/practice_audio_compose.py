# tools/practice_audio_compose.py
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import tempfile
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# =========================
# 数据结构
# =========================

@dataclass
class RowItem:
    no: int
    word_original: str
    audio_primary: str
    timer_ms: int
    combo: int
    sheet: str
    row_idx: int  # excel row index (1-based)


# =========================
# 工具函数
# =========================

def _repo_root() -> str:
    return os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))


def _abs_from_repo(rel_or_abs: str) -> str:
    p = rel_or_abs.strip().strip('"')
    if not p:
        return p
    if os.path.isabs(p):
        return p
    return os.path.abspath(os.path.join(_repo_root(), p))


def _which_or_raise(bin_name: str) -> str:
    # 优先使用 PATH 中的
    try:
        from shutil import which
        p = which(bin_name)
        if p:
            return p
    except Exception:
        pass
    # Windows 常见：ffmpeg/ffprobe 在同目录
    return bin_name


def _run(cmd: List[str]) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="ignore")


def _ffprobe_audio_params(ffprobe_bin: str, mp3_path: str) -> Tuple[int, int]:
    """
    返回 (sample_rate, channels)
    """
    cmd = [
        ffprobe_bin,
        "-v", "error",
        "-select_streams", "a:0",
        "-show_entries", "stream=sample_rate,channels",
        "-of", "json",
        mp3_path,
    ]
    p = _run(cmd)
    if p.returncode == 0:
        try:
            obj = json.loads(p.stdout)
            streams = obj.get("streams", [])
            if streams:
                sr = int(streams[0].get("sample_rate", 22050))
                ch = int(streams[0].get("channels", 1))
                return sr, ch
        except Exception:
            pass
    # fallback
    return 22050, 1

def _ffprobe_duration_ms(ffprobe_bin: str, audio_path: str) -> int:
    """
    返回音频时长（毫秒）。失败则返回 0。
    """
    cmd = [
        ffprobe_bin,
        "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        audio_path,
    ]
    p = _run(cmd)
    if p.returncode != 0:
        return 0
    try:
        sec = float((p.stdout or "").strip())
        return max(0, int(round(sec * 1000)))
    except Exception:
        return 0

def _duration_ms_cached(ffprobe_bin: str, path: str, cache: Dict[str, int]) -> int:
    if path in cache:
        return cache[path]
    ms = _ffprobe_duration_ms(ffprobe_bin, path)
    cache[path] = ms
    return ms

def _ensure_silence_mp3(
    ffmpeg_bin: str,
    silence_dir: str,
    duration_ms: int,
    sample_rate: int,
    channels: int,
    bitrate: str = "96k",
) -> str:
    """
    生成静音 mp3（可复用），返回文件路径
    """
    os.makedirs(silence_dir, exist_ok=True)
    duration_s = max(0, duration_ms) / 1000.0
    # 0ms 不生成文件，返回空
    if duration_s <= 0:
        return ""

    fname = f"silence_{duration_ms}ms_{sample_rate}hz_{channels}ch.mp3"
    out_path = os.path.join(silence_dir, fname)
    if os.path.exists(out_path) and os.path.getsize(out_path) > 0:
        return out_path

    channel_layout = "mono" if channels == 1 else "stereo"
    cmd = [
        ffmpeg_bin,
        "-y",
        "-f", "lavfi",
        "-i", f"anullsrc=r={sample_rate}:cl={channel_layout}",
        "-t", f"{duration_s:.3f}",
        "-q:a", "4",
        "-b:a", bitrate,
        out_path,
    ]
    p = _run(cmd)
    if p.returncode != 0:
        raise RuntimeError(f"ffmpeg failed to create silence mp3: {p.stderr.strip()}")
    return out_path


def _safe_int(v: Any, default: int) -> int:
    try:
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return int(v)
        s = str(v).strip()
        if not s:
            return default
        # 允许 "5000" / "5000.0"
        return int(float(s))
    except Exception:
        return default


def _norm_header(s: str) -> str:
    return re.sub(r"[\s\-]+", "_", (s or "").strip().lower())


def _find_col_index(headers: List[Any], targets: List[str]) -> Optional[int]:
    norm = [_norm_header(str(h)) for h in headers]
    for t in targets:
        t2 = _norm_header(t)
        if t2 in norm:
            return norm.index(t2)
    return None


# =========================
# 读取 Excel / JSON
# =========================

def load_rows_from_excel(xlsx_path: str, only_sheets: Optional[List[str]] = None) -> List[RowItem]:
    wb = load_workbook(xlsx_path, data_only=True)
    rows: List[RowItem] = []

    sheet_names = wb.sheetnames
    if only_sheets:
        sheet_names = [s for s in sheet_names if s in set(only_sheets)]

    for sn in sheet_names:
        ws = wb[sn]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        header = list(all_rows[0])

        idx_no = _find_col_index(header, ["no", "编号", "序号"])
        idx_word = _find_col_index(header, ["word_original", "word", "英文单词", "英文词汇"])
        idx_audio = _find_col_index(header, ["audio_primary", "audio", "audio_url"])
        idx_timer = _find_col_index(header, ["timer", "timer_ms"])
        idx_combo = _find_col_index(header, ["combo", "repeat", "repeat_count"])

        if idx_word is None or idx_audio is None:
            # 这个 sheet 不符合预期，跳过
            continue

        for i, r in enumerate(all_rows[1:], start=2):  # excel row index starts at 1, header at 1
            word = str(r[idx_word] or "").strip()
            audio = str(r[idx_audio] or "").strip()
            if not word or not audio:
                continue

            no = _safe_int(r[idx_no], default=len(rows) + 1) if idx_no is not None else (len(rows) + 1)
            timer_ms = _safe_int(r[idx_timer], default=5000) if idx_timer is not None else 5000
            combo = _safe_int(r[idx_combo], default=2) if idx_combo is not None else 2
            combo = max(1, combo)

            rows.append(RowItem(
                no=no,
                word_original=word,
                audio_primary=audio,
                timer_ms=max(0, timer_ms),
                combo=combo,
                sheet=sn,
                row_idx=i,
            ))

    return rows


def load_rows_from_json(json_path: str) -> List[RowItem]:
    """
    JSON 允许两种格式：
    1) List[dict]，每个 dict 包含：word_original, audio_primary, timer, combo
    2) dict{"items":[...]} 同上
    """
    with open(json_path, "r", encoding="utf-8") as f:
        obj = json.load(f)

    if isinstance(obj, dict) and "items" in obj:
        items = obj["items"]
    else:
        items = obj

    if not isinstance(items, list):
        raise ValueError("JSON must be a list of items or {items:[...]}")

    rows: List[RowItem] = []
    for i, it in enumerate(items, start=1):
        if not isinstance(it, dict):
            continue
        word = str(it.get("word_original", "") or it.get("word", "") or "").strip()
        audio = str(it.get("audio_primary", "") or it.get("audio", "") or "").strip()
        if not word or not audio:
            continue
        timer_ms = _safe_int(it.get("timer", it.get("timer_ms", 5000)), 5000)
        combo = max(1, _safe_int(it.get("combo", 2), 2))
        rows.append(RowItem(
            no=i,
            word_original=word,
            audio_primary=audio,
            timer_ms=max(0, timer_ms),
            combo=combo,
            sheet="JSON",
            row_idx=i,
        ))
    return rows


# =========================
# URL -> 本地文件映射
# =========================
from urllib.parse import urlparse

def url_to_local_audio(audio_url: str, audio_root: str):
    """
    支持：
    - mp3 / wav
    - uk / us
    - URL / 相对路径 / 绝对路径
    """
    s = (audio_url or "").strip()
    if not s:
        return None

    s = s.replace("\\", "/")

    # URL -> path（去掉 query）
    if "://" in s:
        try:
            s = urlparse(s).path
        except Exception:
            pass
    s = s.split("?", 1)[0]

    def pick_existing(p: str):
        if os.path.exists(p) and os.path.getsize(p) > 0:
            return p
        base, ext = os.path.splitext(p)
        if ext.lower() == ".mp3":
            alt = base + ".wav"
            if os.path.exists(alt):
                return alt
        if ext.lower() == ".wav":
            alt = base + ".mp3"
            if os.path.exists(alt):
                return alt
        return None

    # uk / us 优先
    if "/uk/" in s:
        tail = s.split("/uk/", 1)[1]
        return pick_existing(os.path.join(audio_root, "uk", tail))

    if "/us/" in s:
        tail = s.split("/us/", 1)[1]
        return pick_existing(os.path.join(audio_root, "us", tail))

    if s.startswith("uk/") or s.startswith("us/"):
        return pick_existing(os.path.join(audio_root, s))

    # 绝对路径兜底
    if os.path.isabs(s):
        return pick_existing(s)

    return None


# =========================
# 合成主流程
# =========================

def compose_full_mp3(
    ffmpeg_bin: str,
    ffprobe_bin: str,
    rows: List[RowItem],
    audio_root: str,
    output_mp3: str,
    silence_cache_dir: str,
    concat_mode: str = "reencode",
    write_manifest: bool = True,
) -> Tuple[int, int, Optional[str]]:
    """
    返回：(used_items_count, skipped_items_count, manifest_path)
    """
    if not rows:
        raise ValueError("No valid rows to compose.")

    audio_root = _abs_from_repo(audio_root)
    output_mp3 = _abs_from_repo(output_mp3)
    os.makedirs(os.path.dirname(output_mp3), exist_ok=True)

    # 找到第一条可用 mp3 用来探测 sr/ch
    first_local = None
    for r in rows:
        local = url_to_local_audio(r.audio_primary, audio_root)
        if local and os.path.exists(local):
            first_local = local
            break
    if not first_local:
        raise RuntimeError("No audio files found on disk for any row. Check audio_root and audio_primary URLs.")

    sample_rate, channels = _ffprobe_audio_params(ffprobe_bin, first_local)

    # 预生成所需静音片段（按 timer_ms 去重）
    unique_timers = sorted({max(0, r.timer_ms) for r in rows if r.timer_ms > 0})
    silence_map: Dict[int, str] = {}
    for t in unique_timers:
        silence_map[t] = _ensure_silence_mp3(
            ffmpeg_bin=ffmpeg_bin,
            silence_dir=silence_cache_dir,
            duration_ms=t,
            sample_rate=sample_rate,
            channels=channels,
        )

    # 生成 concat 列表
    manifest_items: List[Dict[str, Any]] = []
    timeline_ms = 0

    used = 0
    skipped = 0

    with tempfile.TemporaryDirectory() as td:
        concat_txt = os.path.join(td, "concat_list.txt")

        dur_cache: Dict[str, int] = {}
        with open(concat_txt, "w", encoding="utf-8") as f:
            for r in rows:
                local = url_to_local_audio(r.audio_primary, audio_root)
                if (not local) or (not os.path.exists(local)):
                    skipped += 1
                    continue

                # 每行重复 combo 次
                for k in range(r.combo):
                    # 1️⃣ 当前段开始时间（一定先记）
                    start_ms = timeline_ms

                    # 2️⃣ 计算单词音频真实时长（ms）
                    word_ms = _duration_ms_cached(ffprobe_bin, local, dur_cache)
                    end_ms = start_ms + word_ms

                    # 3️⃣ 记录 manifest（用于前端精确对齐）
                    manifest_items.append({
                        "no": r.no,
                        "word_original": r.word_original,
                        "audio_primary": r.audio_primary,
                        "local_path": local,
                        "start_ms": start_ms,
                        "end_ms": end_ms,
                        "word_duration_ms": word_ms,
                        "sheet": r.sheet,
                        "excel_row": r.row_idx,
                        "repeat_index": k + 1,
                        "timer_ms_after": r.timer_ms,
                    })

                    # 4️⃣ 写入单词音频
                    safe_local = local.replace("'", "'\\''")
                    f.write(f"file '{safe_local}'\n")
                    used += 1

                    # 5️⃣ 写入静音（如果有）
                    sil_ms = 0
                    if r.timer_ms > 0:
                        sil = silence_map.get(r.timer_ms, "")
                        if sil:
                            safe_sil = sil.replace("'", "'\\''")
                            f.write(f"file '{safe_sil}'\n")
                            sil_ms = r.timer_ms

                    # 6️⃣ 推进 timeline（必须在最后）
                    timeline_ms += word_ms + sil_ms

        # 默认重编码，避免不同来源 mp3 在 concat copy 模式下出现爆音/时间轴异常。
        # 如需更快速度，可显式传 concat_mode=copy 或 auto。
        if concat_mode == "copy":
            cmd_copy = [
                ffmpeg_bin, "-y",
                "-f", "concat",
                "-safe", "0",
                "-i", concat_txt,
                "-c", "copy",
                output_mp3,
            ]
            p = _run(cmd_copy)
            if p.returncode != 0:
                raise RuntimeError(f"ffmpeg compose failed in copy mode.\n{p.stderr.strip()}")
        elif concat_mode == "auto":
            # 先尝试 copy，失败再重编码
            cmd_copy = [
                ffmpeg_bin, "-y",
                "-f", "concat",
                "-safe", "0",
                "-i", concat_txt,
                "-c", "copy",
                output_mp3,
            ]
            p = _run(cmd_copy)
            if p.returncode != 0:
                cmd_enc = [
                    ffmpeg_bin, "-y",
                    "-f", "concat",
                    "-safe", "0",
                    "-i", concat_txt,
                    "-c:a", "libmp3lame",
                    "-q:a", "4",
                    output_mp3,
                ]
                p2 = _run(cmd_enc)
                if p2.returncode != 0:
                    raise RuntimeError(
                        "ffmpeg compose failed.\n"
                        f"[copy mode err]\n{p.stderr.strip()}\n\n"
                        f"[re-encode err]\n{p2.stderr.strip()}"
                    )
        else:
            # reencode（默认）
            cmd_enc = [
                ffmpeg_bin, "-y",
                "-f", "concat",
                "-safe", "0",
                "-i", concat_txt,
                "-c:a", "libmp3lame",
                "-q:a", "4",
                output_mp3,
            ]
            p2 = _run(cmd_enc)
            if p2.returncode != 0:
                raise RuntimeError(f"ffmpeg compose failed in re-encode mode.\n{p2.stderr.strip()}")

    manifest_path = None
    if write_manifest:
        manifest_path = os.path.splitext(output_mp3)[0] + ".manifest.json"
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump({
                "output_mp3": output_mp3,
                "audio_root": audio_root,
                "sample_rate": sample_rate,
                "channels": channels,
                "items": manifest_items,
                "note": "start_ms/end_ms are timeline-aligned in ms. timeline accumulates both word audio duration (ffprobe) and timer silence after each repeat.",
            }, f, ensure_ascii=False, indent=2)

    return used, skipped, manifest_path


# =========================
# CLI
# =========================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input_excel", default="", help="path to practice xlsx (preferred)")
    parser.add_argument("--input_json", default="", help="optional path to json list of items")
    parser.add_argument("--audio_root", default="storage/audio_cache", help="where uk/us mp3 live")
    parser.add_argument("--output_mp3", required=True, help="output mp3 file path")

    parser.add_argument("--ffmpeg_bin", default="ffmpeg", help="ffmpeg executable name or path")
    parser.add_argument("--ffprobe_bin", default="ffprobe", help="ffprobe executable name or path")

    parser.add_argument("--only_sheets", default="", help="comma-separated sheet names; empty=all")
    parser.add_argument("--silence_cache_dir", default="storage/audio_cache/_silence",
                        help="cache dir for silence mp3 segments")
    parser.add_argument(
        "--concat_mode",
        default="reencode",
        choices=["reencode", "copy", "auto"],
        help="concat strategy: reencode(default, most stable), copy(fast but fragile), auto(copy then fallback reencode)",
    )
    parser.add_argument("--no_manifest", action="store_true", help="do not write manifest json")
    args = parser.parse_args()

    if not args.input_excel and not args.input_json:
        raise SystemExit("Provide --input_excel or --input_json")

    ffmpeg_bin = _which_or_raise(args.ffmpeg_bin)
    ffprobe_bin = _which_or_raise(args.ffprobe_bin)

    rows: List[RowItem] = []
    if args.input_excel:
        only_sheets = [s.strip() for s in args.only_sheets.split(",") if s.strip()] if args.only_sheets else None
        rows = load_rows_from_excel(_abs_from_repo(args.input_excel), only_sheets=only_sheets)
    else:
        rows = load_rows_from_json(_abs_from_repo(args.input_json))

    if not rows:
        raise SystemExit("No valid rows found. Check columns: word_original, audio_primary, timer, combo")

    used, skipped, manifest_path = compose_full_mp3(
        ffmpeg_bin=ffmpeg_bin,
        ffprobe_bin=ffprobe_bin,
        rows=rows,
        audio_root=args.audio_root,
        output_mp3=args.output_mp3,
        silence_cache_dir=_abs_from_repo(args.silence_cache_dir),
        concat_mode=args.concat_mode,
        write_manifest=(not args.no_manifest),
    )

    print(f"[DONE] composed_mp3={_abs_from_repo(args.output_mp3)} used_segments={used} skipped_rows={skipped}")
    if manifest_path:
        print(f"[DONE] manifest={manifest_path}")


if __name__ == "__main__":
    main()
