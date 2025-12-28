# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple

import openpyxl
from fastapi import APIRouter, HTTPException, Request

from api.schemas import (
    AudioComposeRequest, AudioComposeResponse,
    VocabNoteAudioRequest, VocabNoteAudioResponse,
    FillForVocabNoteRequest, FillForVocabNoteResponse,
    GenerateFromVocabNoteRequest, GenerateFromVocabNoteResponse
)

from utils.slug import safe_filename_from_word

router = APIRouter(tags=["audio"])

# routes_audio.py
try:
    from utils.slug import normalize_word as _normalize_word  # 与 audio_cache_build.py 同源
except Exception:
    _normalize_word = None

def _norm_word(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    if _normalize_word:
        return _normalize_word(s).lower()
    # fallback：至少做“空白归一 + 小写”
    s = re.sub(r"\s+", " ", s)
    return s.lower()

def _ensure_parent(p):
    Path(p).parent.mkdir(parents=True, exist_ok=True)

def _project_root() -> Path:
    # api/routes_audio.py -> api -> vocab_service(root)
    return Path(__file__).resolve().parents[1]


def _list_outputs(out_dir: Path) -> List[str]:
    if not out_dir.exists():
        return []
    files = []
    for p in out_dir.rglob("*"):
        if p.is_file() and p.suffix.lower() in (".mp3", ".json"):
            files.append(str(p.resolve()))
    files.sort()
    return files


def _env_required(k: str) -> str:
    v = (os.getenv(k) or "").strip()
    if not v:
        raise RuntimeError(f"Missing env: {k}")
    return v


def _read_words_from_excel(vocab_excel: Path, sheet_name: str) -> List[str]:
    wb = openpyxl.load_workbook(vocab_excel, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    header = [(str(c.value).strip() if c.value is not None else "") for c in ws[1]]
    header_l = [h.lower() for h in header]

    # 兼容：word_original / word / word_norm / 中文表头
    candidates = ["word_original", "word", "word_norm", "英文单词", "单词", "词汇", "英文", "english", "english word"]
    col_idx = None
    for cand in candidates:
        if cand.lower() in header_l:
            col_idx = header_l.index(cand.lower()) + 1
            break
    if col_idx is None:
        raise HTTPException(status_code=400, detail=f"vocab_note.xlsx missing word column. header={header}")

    def _norm(s: str) -> str:
        return _norm_word(s)

    out: List[str] = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None:
            continue
        w = _norm(str(v))
        if w:
            out.append(w)

    # 去重保持顺序
    seen: Set[str] = set()
    dedup: List[str] = []
    for w in out:
        if w not in seen:
            seen.add(w)
            dedup.append(w)
    return dedup


def _load_json_any(p: Path) -> Any:
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


def _flatten_cache_entries(cache_obj: Any) -> List[Dict[str, Any]]:
    """
    兼容：
    A) {"entries":[...]}
    B) [...]
    C) {"word": {"entries":[...]}, ...}
    """
    out: List[Dict[str, Any]] = []
    if cache_obj is None:
        return out

    if isinstance(cache_obj, dict) and isinstance(cache_obj.get("entries"), list):
        for e in cache_obj["entries"]:
            if isinstance(e, dict):
                out.append(e)
        return out

    if isinstance(cache_obj, list):
        for e in cache_obj:
            if isinstance(e, dict):
                out.append(e)
        return out

    if isinstance(cache_obj, dict):
        for _, v in cache_obj.items():
            if not isinstance(v, dict):
                continue
            lst = v.get("entries")
            if isinstance(lst, list):
                for e in lst:
                    if isinstance(e, dict):
                        out.append(e)
            else:
                if "word_original" in v or "word_norm" in v or "word" in v:
                    out.append(v)
    return out


def _build_audio_index(global_entries: List[Dict[str, Any]], uploaded_entries: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    def _k(e: Dict[str, Any]) -> str:
        return _norm_word(str(e.get("word_original") or e.get("word") or e.get("word_norm") or ""))

    idx: Dict[str, Dict[str, Any]] = {}
    for e in global_entries:
        key = _k(e)
        if key and key not in idx:
            idx[key] = e
    for e in uploaded_entries:
        key = _k(e)
        if key:
            idx[key] = e
    return idx


def _ref_to_disk(audio_root: Path, ref: str) -> Path:
    """
    支持 /static/audio/uk/xxx.mp3 这种 ref
    """
    s = (ref or "").strip()
    if not s:
        return Path("")
    m = re.search(r"/(uk|us)/([^/]+\.(mp3|wav))$", s, re.IGNORECASE)
    if m:
        return (audio_root / m.group(1).lower() / m.group(2)).resolve()
    return (audio_root / s.lstrip("/\\")).resolve()


def _probe_missing(words: List[str], idx: Dict[str, Dict[str, Any]], accents: List[str], audio_root: Path) -> List[str]:
    missing: List[str] = []
    for w in words:
        e = idx.get(w)
        if not e:
            missing.append(w)
            continue
        ok = True
        for ac in accents:
            ac = ac.lower()
            field = "audio_uk" if ac == "uk" else "audio_us"
            ref = (e.get(field) or "").strip()
            if not ref:
                ok = False
                break
            if not _ref_to_disk(audio_root, ref).exists():
                ok = False
                break
        if not ok:
            missing.append(w)
    return missing


def _append_queue(queue_path: Path, missing_words: List[str], accents: List[str]) -> None:
    queue_path.parent.mkdir(parents=True, exist_ok=True)
    with queue_path.open("a", encoding="utf-8") as f:
        for w in missing_words:
            for ac in accents:
                f.write(json.dumps({"word_norm": w, "accent": ac}, ensure_ascii=False) + "\n")


@router.post(
    "/fill_for_vocab_note",
    response_model=FillForVocabNoteResponse,
    tags=["DEPRECATED_DO_NOT_TOUCH"],
    deprecated=True,
    include_in_schema=False,   # ✅ Swagger 隐藏
)
def fill_for_vocab_note(req: FillForVocabNoteRequest):
    """
    ⚠️ DEPRECATED / LEGACY ENDPOINT — DO NOT TOUCH ⚠️

    状态说明：
    - 该接口为历史遗留接口，仅用于兼容旧调用
    - 已被新链路替代：
        POST /v1/audio/generate_from_vocab_note
        → POST /v1/audio/vocab_note
    - ⚠️ 请勿修改 / 请勿重构 / 请勿作为新功能入口

    给未来自己的备注（2025-01）：
    半个月后再看到这里，直接跳过，不要动它。
    """

    ROOT = _project_root()
    # ↓↓↓ 原有代码一律不动 ↓↓↓
    try:
        vocab_excel = Path(req.vocab_excel)
        if not vocab_excel.is_absolute():
            vocab_excel = (ROOT / vocab_excel).resolve()
        if not vocab_excel.exists():
            raise HTTPException(status_code=400, detail=f"vocab_excel not found: {vocab_excel}")

        # env
        global_cache = Path(_env_required("GLOBAL_CACHE"))
        uploaded_cache = Path(_env_required("UPLOADED_CACHE"))
        audio_root = Path(_env_required("AUDIO_ROOT"))
        queue_path = Path(os.getenv("MISSING_AUDIO_QUEUE") or "storage/missing_audio_queue.jsonl")
        if not queue_path.is_absolute():
            queue_path = (ROOT / queue_path).resolve()

        # 读本次词
        words = _read_words_from_excel(vocab_excel, req.sheet_name)

        # 建索引（uploaded 覆盖 global）
        gobj = _load_json_any((ROOT / global_cache).resolve() if not global_cache.is_absolute() else global_cache)
        uobj = _load_json_any((ROOT / uploaded_cache).resolve() if not uploaded_cache.is_absolute() else uploaded_cache)
        g_entries = _flatten_cache_entries(gobj)
        u_entries = _flatten_cache_entries(uobj)
        idx = _build_audio_index(g_entries, u_entries)

        accents = [a.lower() for a in (req.accents or ["uk", "us"]) if a.lower() in ("uk", "us")]
        if not accents:
            accents = ["uk", "us"]

        missing = _probe_missing(words, idx, accents, audio_root.resolve() if audio_root.is_absolute() else (ROOT / audio_root).resolve())
        if not missing:
            return FillForVocabNoteResponse(
                run_id=req.run_id or "",
                vocab_excel=str(vocab_excel),
                sheet_name=req.sheet_name,
                accents=accents,
                missing_count=0,
                missing_sample=[],
                processed_words=0,
                queue_path=str(queue_path),
                detail="audio already complete; nothing to fill",
            )

        # 限制本次处理量
        to_fill = missing[: max(1, int(req.max_words or 200))]

        # 入队
        _append_queue(queue_path, to_fill, accents)

        # 立刻同步跑一批 fill（本地用最舒服；上线也可以先这样，之后换后台 worker）
        piper_bin = _env_required("PIPER_BIN")
        uk_model = _env_required("PIPER_UK_MODEL")
        us_model = _env_required("PIPER_US_MODEL")
        url_prefix = (os.getenv("AUDIO_URL_PREFIX") or "/static/audio").strip()

        audio_missing_fill_py = (ROOT / "tools" / "audio_missing_fill.py").resolve()
        if not audio_missing_fill_py.exists():
            raise HTTPException(status_code=500, detail=f"audio_missing_fill.py not found: {audio_missing_fill_py}")

        cmd = [
            sys.executable, str(audio_missing_fill_py),
            "--queue_path", str(queue_path),
            "--global_cache", str(global_cache),
            "--uploaded_cache", str(uploaded_cache),
            "--audio_root", str(audio_root),
            "--format", "mp3",
            "--piper_bin", piper_bin,
            "--uk_model", uk_model,
            "--us_model", us_model,
            "--url_mode", "static",
            "--url_prefix", url_prefix,
            "--batch", str(len(to_fill) * len(accents)),
        ]
        if req.write_uploaded:
            cmd.append("--write_uploaded")

        ret = subprocess.run(cmd, cwd=str(ROOT), capture_output=False)
        if ret.returncode != 0:
            raise HTTPException(status_code=500, detail=f"fill failed, returncode={ret.returncode}")

        return FillForVocabNoteResponse(
            run_id=req.run_id or "",
            vocab_excel=str(vocab_excel),
            sheet_name=req.sheet_name,
            accents=accents,
            missing_count=len(missing),
            missing_sample=missing[:20],
            processed_words=len(to_fill),
            queue_path=str(queue_path),
            detail="enqueued and filled a batch; re-run compose after this",
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"[N/A] /audio/fill_for_vocab_note error: {e}")

# 合成 practice的乱序音频

@router.post("/compose", response_model=AudioComposeResponse)
def compose_audio(req: AudioComposeRequest, request: Request):
    """
    镜像 export/practice：
    - 输入：master_xlsx（practice_master.xlsx）
    - 输出：storage/out/{run_id}/audio/*.mp3 + *.manifest.json
    - 同目录写 audio.meta.json
    - 调用 tools/audio_orchestrator.py（它会再调用 tools/practice_audio_compose.py）
    """
    rid = request.headers.get("x-request-id") or "N/A"

    run_id = (req.run_id or "").strip()
    if not run_id:
        raise HTTPException(status_code=400, detail="run_id is required")

    master_xlsx = (req.master_xlsx or "").strip()
    if not master_xlsx:
        raise HTTPException(status_code=400, detail="master_xlsx is required")

    master_path = Path(master_xlsx).expanduser()
    if not master_path.exists():
        raise HTTPException(status_code=400, detail=f"master_xlsx not found: {master_path}")

    ROOT = _project_root()

    # 默认输出目录：storage/out/{run_id}/audio
    out_dir = (req.output_dir or "").strip()
    if not out_dir:
        out_dir = str((ROOT / "storage" / "out" / run_id / "audio").resolve())
    out_dir_path = Path(out_dir)
    out_dir_path.mkdir(parents=True, exist_ok=True)

    # tools 脚本路径（你说都在 tools 目录）
    orchestrator = (ROOT / "tools" / "audio_orchestrator.py").resolve()
    if not orchestrator.exists():
        raise HTTPException(status_code=500, detail=f"audio_orchestrator.py not found: {orchestrator}")

    # ====== PRECHECK：确保音频已补齐，否则拒绝 compose ======
    audio_root = (req.audio_root or "").strip() or _env_required("AUDIO_ROOT")
    audio_root_path = Path(audio_root)
    if not audio_root_path.is_absolute():
        audio_root_path = (ROOT / audio_root_path).resolve()

    # 这里需要一个 probe_audio_available(master_path, audio_root_path)
    # 如果 missing > 0：raise HTTPException(400, detail=...)

    # 构建命令：python tools/audio_orchestrator.py --input_excel ... --output_dir ...
    cmd = [
        sys.executable,
        str(orchestrator),
        "--input_excel", str(master_path),
        "--output_dir", str(out_dir_path),
    ]

    # 可选参数
    sheet = (req.sheet or "").strip()
    if sheet:
        cmd += ["--sheet", sheet]

    prefix = (req.prefix or "").strip()
    if prefix:
        cmd += ["--prefix", prefix]

    if req.make_parts:
        cmd += ["--make_parts"]

    audio_root = (req.audio_root or "").strip()
    if audio_root:
        cmd += ["--audio_root", audio_root]

    ffmpeg_bin = (req.ffmpeg_bin or "").strip()
    if ffmpeg_bin:
        cmd += ["--ffmpeg_bin", ffmpeg_bin]

    ffprobe_bin = (req.ffprobe_bin or "").strip()
    if ffprobe_bin:
        cmd += ["--ffprobe_bin", ffprobe_bin]

    if req.no_manifest:
        cmd += ["--no_manifest"]

    try:
        # 执行脚本
        proc = subprocess.run(
            cmd,
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            check=False,
        )
        if proc.returncode != 0:
            # 把 stderr 末尾带回去方便你定位
            tail = (proc.stderr or proc.stdout or "").strip()[-1200:]
            raise HTTPException(status_code=500, detail=f"[{rid}] audio compose failed (code={proc.returncode}): {tail}")

        # 扫描输出目录产物
        generated = _list_outputs(out_dir_path)

        # 写 meta.json（镜像 export/practice）
        meta_path = (out_dir_path / "audio.meta.json").resolve()
        meta = {
            "run_id": run_id,
            "master_xlsx": str(master_path),
            "output_dir": str(out_dir_path.resolve()),
            "cmd": cmd,
            "generated": generated,
        }
        meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

        return AudioComposeResponse(
            run_id=run_id,
            master_xlsx=str(master_path),
            output_dir=str(out_dir_path.resolve()),
            generated=generated,
            meta_json=str(meta_path),
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"[{rid}] audio compose failed: {e}")

# vocab_note 正序词汇音频合成

@router.post("/vocab_note", response_model=VocabNoteAudioResponse)
def compose_vocab_note_audio(req: VocabNoteAudioRequest, request: Request):
    """
    vocab_note 正序词汇音频：
    - vocab_note.xlsx 不包含 audio_primary 时：
      用 word_original 去匹配 global_cache + uploaded_vocab_cache，拼 input_json 再 compose
    - 默认会生成 manifest.json（除非 no_manifest=True）
    """
    rid = request.headers.get("x-request-id") or "N/A"
    try:
        ROOT = _project_root()

        run_id = (req.run_id or "").strip()
        if not run_id:
            raise HTTPException(status_code=400, detail="run_id is required")

        vocab_excel = (req.vocab_excel or "").strip()
        if not vocab_excel:
            raise HTTPException(status_code=400, detail="vocab_excel is required (vocab_note.xlsx)")
        if not os.path.exists(vocab_excel):
            raise HTTPException(status_code=400, detail=f"vocab_excel not found: {vocab_excel}")

        audio_root = (req.audio_root or "storage/audio_cache").strip()
        accent = (req.accent or "uk").strip().lower()
        if accent not in ("uk", "us"):
            raise HTTPException(status_code=400, detail="accent must be 'uk' or 'us'")

        out_mp3 = (req.output_mp3 or "").strip()
        if not out_mp3:
            out_dir = ROOT / "storage" / "out" / run_id / "audio"
            out_dir.mkdir(parents=True, exist_ok=True)
            out_mp3 = str((out_dir / f"vocab_note__{run_id}.mp3").resolve())
        _ensure_parent(out_mp3)

        compose_py = (ROOT / "tools" / "practice_audio_compose.py").resolve()

        # ---------- 1) 读取 vocab_note.xlsx ----------
        from openpyxl import load_workbook

        wb = load_workbook(vocab_excel, data_only=True)
        sheet_name = (req.only_sheets or "").strip()
        if sheet_name and sheet_name in wb.sheetnames:
            sheets = [sheet_name]
        else:
            sheets = [wb.sheetnames[0]]

        # 只取第一个 sheet（vocab_note 正序一般就一个）
        ws = wb[sheets[0]]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise HTTPException(status_code=400, detail="vocab_note sheet is empty")

        header = [str(x).strip() if x is not None else "" for x in header_row]
        header_l = [h.lower() for h in header]

        # 必须有 word_original
        def _find_col(header_l, candidates):
            for c in candidates:
                c = c.lower()
                if c in header_l:
                    return header_l.index(c)
            return None

        # ✅ 兼容 vocab_note 可能出现的列名
        WORD_COL_CANDIDATES = [
            "word_original",  # 你期望的 key
            "word",           # 常见英文
            "英文单词", "单词", "词汇", "词", "英文", "word (english)", "english word"
        ]

        w_idx = _find_col(header_l, WORD_COL_CANDIDATES)
        if w_idx is None:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "vocab_note_missing_word_column",
                    "expected_any_of": WORD_COL_CANDIDATES,
                    "actual_header": header,
                },
            )

        # 可选：如果未来你 export 加了 audio_primary，可直接用
        AUDIO_COL_CANDIDATES = [
            "audio_primary", "audio", "audio_url",
            "音频", "音频链接", "发音", "发音链接",
            "audio(primary)", "audio link"
        ]
        a_idx = _find_col(header_l, AUDIO_COL_CANDIDATES)

        words_in_order = []
        excel_audio_map = {}  # word_norm -> audio_url from excel (if present)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if w_idx >= len(row):
                continue
            w = row[w_idx]
            if w is None:
                continue
            w = str(w).strip()
            if not w:
                continue
            wn = _norm_word(w)
            words_in_order.append((w, wn))
            if a_idx is not None and a_idx < len(row):
                av = row[a_idx]
                if av is not None and str(av).strip():
                    excel_audio_map[wn] = str(av).strip()

        if not words_in_order:
            raise HTTPException(status_code=400, detail="vocab_note.xlsx has no words")

        # ---------- 2) 加载 caches 并建立索引 ----------
        def _load_cache_any(p: str) -> object:
            pp = Path(p)
            if not pp.is_absolute():
                pp = (ROOT / pp).resolve()
            if not pp.exists():
                return {}
            return json.loads(pp.read_text(encoding="utf-8"))

        def _flatten_cache_entries(cache_obj: object) -> list[dict]:
            """
            兼容三种 cache 形态：
            A) {"entries":[{...},{...}]}
            B) [{...},{...}]
            C) {"organic farming": {"entries":[{...}]}, "durable": {"entries":[{...}]}, ...}
            """
            out: list[dict] = []

            if cache_obj is None:
                return out

            # A) {"entries":[...]}
            if isinstance(cache_obj, dict) and "entries" in cache_obj and isinstance(cache_obj["entries"], list):
                for e in cache_obj["entries"]:
                    if isinstance(e, dict):
                        out.append(e)
                return out

            # B) [...]
            if isinstance(cache_obj, list):
                for e in cache_obj:
                    if isinstance(e, dict):
                        out.append(e)
                return out

            # C) {"word": {"entries":[...]}, ...}
            if isinstance(cache_obj, dict):
                for _, v in cache_obj.items():
                    if not isinstance(v, dict):
                        continue

                    lst = v.get("entries")
                    if isinstance(lst, list):
                        for e in lst:
                            if isinstance(e, dict):
                                out.append(e)
                    else:
                        # 兜底：有些结构可能把 entry 直接放在 v 里
                        if isinstance(v, dict) and ("word_original" in v or "word" in v or "word_norm" in v):
                            out.append(v)

            return out
        # ✅ 优先用 env（与 fill_for_vocab_note 一致），请求没传就走 env
        global_obj = _load_cache_any(req.global_cache or _env_required("GLOBAL_CACHE"))
        uploaded_obj = _load_cache_any(req.uploaded_cache or _env_required("UPLOADED_CACHE"))

        global_entries = _flatten_cache_entries(global_obj)
        uploaded_entries = _flatten_cache_entries(uploaded_obj)

        def _resolve_audio_ref(
            *,
            word_original: str,
            word_norm: str,
            accent: str,
            excel_audio_map: dict,
            idx: dict,
            audio_root: Path,
        ):
            """
            返回一个“确定可用”的 audio_url，或 None
            """
            # 1) excel 自带
            audio_url = excel_audio_map.get(word_norm)
            if audio_url:
                try:
                    if _ref_to_disk(audio_root, audio_url).exists():
                        return audio_url
                except Exception:
                    pass

            # 2) cache
            e = idx.get(word_norm)
            if e:
                audio_url = _pick_audio(e)
                if audio_url:
                    try:
                        if _ref_to_disk(audio_root, audio_url).exists():
                            return audio_url
                    except Exception:
                        pass

            # 3) 磁盘兜底
            slug = safe_filename_from_word(word_original)
            for ext in (".mp3", ".wav"):
                p = audio_root / accent / f"{slug}{ext}"
                if p.exists():
                    return f"{os.getenv('AUDIO_URL_PREFIX', '/static/audio')}/{accent}/{p.name}"

            return None

        def _pick_audio(entry: dict) -> str:
            ap = (entry.get("audio_primary") or "").strip()
            if ap:
                return ap

            if accent == "uk":
                v = (entry.get("audio_uk") or entry.get("audio") or "").strip()
                if v:
                    return v
            else:
                v = (entry.get("audio_us") or entry.get("audio") or "").strip()
                if v:
                    return v

            for k in ("audio_uk", "audio_us", "audio"):
                v = (entry.get(k) or "").strip()
                if v:
                    return v
            return ""

        idx: dict[str, dict] = {}

        # global 先写
        for e in global_entries:
            if isinstance(e, dict):
                key = _norm_word(e.get("word_original") or e.get("word") or e.get("word_norm") or "")
                if key and key not in idx:
                    idx[key] = e
        # uploaded 覆盖（用户上传优先）
        for e in uploaded_entries:
            if isinstance(e, dict):
                key = _norm_word(e.get("word_original") or e.get("word") or e.get("word_norm") or "")
                if key:
                    idx[key] = e
        # ============================================================
        # ✅ 强制前置 precheck：优先认“磁盘真实文件”
        #    规则：
        #    1) 如果 excel 自带 audio_url 且 _ref_to_disk 存在 -> 通过
        #    2) 否则如果 cache 有 audio_url 且 _ref_to_disk 存在 -> 通过
        #    3) 否则兜底：按 safe_filename(word_original) 去 AUDIO_ROOT/acc/ 查 mp3/wav -> 存在则通过
        #    4) 以上都不满足 -> missing
        # ============================================================
        missing = []

        # accent：你后面合成使用哪个口音，就用哪个口音做 precheck
        acc = (req.accent or "uk").lower().strip()
        if acc not in ("uk", "us"):
            acc = "uk"

        # audio_root：优先用请求传的，否则用 env 的 AUDIO_ROOT（和你系统一致）
        audio_root = Path(req.audio_root or _env_required("AUDIO_ROOT")).resolve()

        for w, wn in words_in_order:
            ok = False

            # 1) excel 自带 audio_url
            audio_url = excel_audio_map.get(wn, "") or ""
            if audio_url:
                try:
                    p = _ref_to_disk(audio_url, audio_root)
                    if p.exists():
                        ok = True
                except Exception:
                    pass

            # 2) cache 的 audio_url
            if not ok:
                e = idx.get(wn)
                if e:
                    try:
                        audio_url2 = _pick_audio(e) or ""
                        if audio_url2:
                            p = _ref_to_disk(audio_url2, audio_root)
                            if p.exists():
                                ok = True
                    except Exception:
                        pass

            # 3) 磁盘兜底：safe_filename(word_original)
            if not ok:
                slug = safe_filename_from_word(w)
                p_mp3 = audio_root / acc / f"{slug}.mp3"
                p_wav = audio_root / acc / f"{slug}.wav"
                if p_mp3.exists() or p_wav.exists():
                    ok = True

            if not ok:
                missing.append(w)

        if missing:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "audio_missing_precheck",
                    "missing_count": len(missing),
                    "missing_sample": missing[:20],
                    "hint": "Call POST /v1/audio/generate_from_vocab_note first (or /v1/audio/fill_for_vocab_note), then retry /v1/audio/vocab_note."
                }
            )

        # ============================================================
        # ---------- 3) 生成 input_json items（保持 vocab_note 顺序） ----------
        # （到这里 missing 已经保证为 0 了）
        # ============================================================
        items = []
        for w, wn in words_in_order:
            # 优先用 excel 的 audio（如果存在）
            audio_url = _resolve_audio_ref(
                word_original=w,
                word_norm=wn,
                accent=accent,              # 注意：这里用你上面已经校验过的 accent 变量
                excel_audio_map=excel_audio_map,
                idx=idx,
                audio_root=Path(audio_root).resolve() if not isinstance(audio_root, Path) else audio_root,
            )

            if not audio_url:
                raise HTTPException(
                    status_code=500,
                    detail=f"precheck passed but audio still missing for: {w}"
                )
            
            # ✅ 这里保持你原来的 item 结构怎么写就怎么写
            items.append({
                "word_original": w,
                "audio_primary": audio_url,
                "timer": 2000,   # ✅ 改这里：你想要的间隔（毫秒）
                "combo": 2       # ✅ 保持默认重复2次（可留可不留）
            })

        if not items:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "no_audio_mapped_for_vocab_note",
                    "missing_count": len(missing),
                    "missing_sample": missing[:20],
                    "hint": "Ensure caches contain audio_uk/audio_us/audio_primary for these words, or run your audio_cache_build pipeline first.",
                }
            )

        # 如果少量缺失，不阻断，但给你提示（你也可以改成严格模式）
        # 这里不抛错，继续生成。
        # ---------- 4) 写临时 json 并调用 compose ----------
        out_dir = Path(out_mp3).parent
        tmp_json = (out_dir / f"_vocab_note_audio_input__{run_id}.json").resolve()
        tmp_json.write_text(json.dumps({"items": items}, ensure_ascii=False, indent=2), encoding="utf-8")

        cmd = [
            sys.executable,
            str(compose_py),
            "--input_json", str(tmp_json),
            "--audio_root", str(audio_root),
            "--output_mp3", out_mp3,
            "--ffmpeg_bin", (req.ffmpeg_bin or "ffmpeg"),
            "--ffprobe_bin", (req.ffprobe_bin or "ffprobe"),
        ]
        if req.no_manifest:
            cmd += ["--no_manifest"]

        p = subprocess.run(cmd, cwd=str(ROOT), text=True, capture_output=True)
        if p.returncode != 0:
            tail = (p.stderr or p.stdout or "").strip()[-1500:]
            raise HTTPException(status_code=500, detail=f"[{rid}] vocab_note compose failed: {tail}")

        manifest = os.path.splitext(out_mp3)[0] + ".manifest.json"

        # 写一个小 meta，方便你排查（可选但建议）
        meta_path = (out_dir / f"vocab_note_audio.meta.json").resolve()
        meta_path.write_text(
            json.dumps(
                {
                    "run_id": run_id,
                    "vocab_excel": vocab_excel,
                    "audio_root": str(Path(audio_root).resolve()),
                    "accent": accent,
                    "items_count": len(items),
                    "missing_count": len(missing),
                    "missing_sample": missing[:50],
                    "tmp_json": str(tmp_json),
                    "output_mp3": out_mp3,
                    "manifest": manifest if (not req.no_manifest and os.path.exists(manifest)) else "",
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        return VocabNoteAudioResponse(
            run_id=run_id,
            vocab_excel=vocab_excel,
            output_mp3=out_mp3,
            manifest_json=manifest if (not req.no_manifest and os.path.exists(manifest)) else "",
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"[{rid}] /audio/vocab_note error: {e}")


@router.post("/generate_from_vocab_note")
def generate_from_vocab_note(req: GenerateFromVocabNoteRequest):
    ROOT = _project_root()

    tool = (ROOT / "tools" / "audio_generate_from_vocab_note.py").resolve()
    if not tool.exists():
        raise HTTPException(status_code=500, detail=f"tool not found: {tool}")

    payload_path = (ROOT / "storage" / "tmp" / "gen_audio_payload.json").resolve()
    payload_path.parent.mkdir(parents=True, exist_ok=True)
    payload_path.write_text(
        json.dumps(req.model_dump(), ensure_ascii=False),
        encoding="utf-8"
    )

    cmd = [sys.executable, str(tool), str(payload_path)]
    p = subprocess.run(cmd, cwd=str(ROOT), capture_output=True, text=True)

    if p.returncode != 0:
        tail = (p.stderr or p.stdout or "").strip()[-1500:]
        raise HTTPException(status_code=500, detail=f"generate_from_vocab_note failed: {tail}")

    try:
        return json.loads(p.stdout)
    except Exception:
        raise HTTPException(status_code=500, detail=f"tool output is not json: {p.stdout[-1500:]}")

