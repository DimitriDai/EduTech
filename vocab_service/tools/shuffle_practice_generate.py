# tools/shuffle_practice_generate.py
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import os
import random
import sys
import unicodedata
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# =========================
# 路径与工具
# =========================

def repo_root() -> str:
    return os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))


def try_import_normalize():
    """
    尽量复用你项目里的 normalize/slug 规则（稳定性更高）。
    """
    sys.path.insert(0, repo_root())
    try:
        from utils.slug import normalize_word  # type: ignore
        return normalize_word
    except Exception:
        return None


def fallback_normalize_word(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKC", s)
    s = " ".join(s.split())
    return s


def load_json(path: str) -> Dict[str, Any]:
    if not path or not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def iter_entries(cache_obj: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    兼容你的 cache 结构：top_key -> { word_norm, entries:[...] }
    扁平化返回 entries 列表。
    """
    out: List[Dict[str, Any]] = []
    for _, group in (cache_obj or {}).items():
        if not isinstance(group, dict):
            continue
        entries = group.get("entries", [])
        if isinstance(entries, list):
            for e in entries:
                if isinstance(e, dict):
                    out.append(e)
    return out


def resolve_audio_primary(e: Dict[str, Any], prefer: str = "uk") -> str:
    """
    audio_primary 优先：entry.audio_primary
    否则 fallback: prefer(uk/us) -> other
    """
    ap = str(e.get("audio_primary", "") or "").strip()
    if ap:
        return ap

    uk = str(e.get("audio_uk", "") or "").strip()
    us = str(e.get("audio_us", "") or "").strip()

    if prefer == "uk":
        return uk or us
    return us or uk


def resolve_pos_cn(e: Dict[str, Any]) -> str:
    """
    你的 schema 里一般是 pos_cn（词性+中文释义）
    如果没有就尽量 fallback。
    """
    v = e.get("pos_cn", None)
    if isinstance(v, str) and v.strip():
        return v.strip()

    # 兼容可能存在的 meaning/definition_cn 等字段名
    for k in ("meaning", "definition_cn", "cn", "pos+cn"):
        vv = e.get(k, None)
        if isinstance(vv, str) and vv.strip():
            return vv.strip()

    return ""


def choose_best_entry_per_word(entries: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    同一个 word_norm 可能有多个 entry（不同来源/不同字段完整度）。
    这里取“最完整”的一个：按字段非空数排序。
    """
    buckets: Dict[str, List[Dict[str, Any]]] = {}
    for e in entries:
        wn = str(e.get("word_norm", "") or "").strip()
        if not wn:
            continue
        buckets.setdefault(wn, []).append(e)

    def score(e: Dict[str, Any]) -> int:
        keys = ["phonetic_uk", "phonetic_us", "pos_cn", "definition_en", "example", "example_cn",
                "synonyms", "audio_primary", "audio_uk", "audio_us"]
        s = 0
        for k in keys:
            v = e.get(k, "")
            if isinstance(v, str):
                if v.strip():
                    s += 1
            elif v is not None:
                s += 1
        return s

    picked: List[Dict[str, Any]] = []
    for wn, lst in buckets.items():
        lst_sorted = sorted(lst, key=score, reverse=True)
        picked.append(lst_sorted[0])
    return picked


# =========================
# Excel 写入
# =========================

def set_column_widths(ws, widths: List[int]) -> None:
    for idx, w in enumerate(widths, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = w


def build_rows(
    entries: List[Dict[str, Any]],
    prefer_audio: str,
    timer_ms: int,
    combo_count: int,
    require_audio: bool,
) -> List[List[Any]]:
    rows: List[List[Any]] = []
    no = 1
    for e in entries:
        word = str(e.get("word_original", "") or "").strip()
        if not word:
            continue

        audio_primary = resolve_audio_primary(e, prefer=prefer_audio).strip()
        if require_audio and (not audio_primary):
            continue

        pos_cn = resolve_pos_cn(e)

        # 练习表（英译中）推荐列：
        # No | word_original | 学生答案 | pos_cn(中文答案) | audio_primary | timer | combo
        rows.append([
            no,
            word,
            "",  # 学生答案留空
            pos_cn,
            audio_primary,
            int(timer_ms),
            int(combo_count),
        ])
        no += 1

    return rows


def write_workbook(
    rows: List[List[Any]],
    output_xlsx: str,
    sheet_size: int = 25,
) -> None:
    wb = Workbook()
    # 删除默认 sheet，自己创建更可控
    wb.remove(wb.active)

    headers = ["no", "word_original", "student_answer", "pos_cn", "audio_primary", "timer", "combo"]

    total = len(rows)
    if total == 0:
        # 至少输出一个空 sheet + 表头
        ws = wb.create_sheet("Sheet1")
        ws.append(headers)
        set_column_widths(ws, [6, 22, 18, 38, 38, 10, 10])
        os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
        wb.save(output_xlsx)
        return

    sheet_count = (total + sheet_size - 1) // sheet_size

    for i in range(sheet_count):
        start = i * sheet_size
        end = min((i + 1) * sheet_size, total)
        ws = wb.create_sheet(f"Sheet{i+1}")
        ws.append(headers)
        for r in rows[start:end]:
            ws.append(r)

        # 宽度你可后续微调
        set_column_widths(ws, [6, 22, 18, 38, 38, 10, 10])

    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    wb.save(output_xlsx)


# =========================
# 主流程
# =========================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--global_cache", default="storage/global_cache.json")
    parser.add_argument("--uploaded_cache", default="storage/uploaded_vocab_cache.json")
    parser.add_argument("--source", choices=["global", "uploaded", "both"], default="uploaded",
                        help="乱序练习一般用 uploaded 作为大词库来源；global 用于你精选词")
    parser.add_argument("--output", required=True, help="output xlsx path")
    parser.add_argument("--limit", type=int, default=500, help="max entries to output (after dedupe/filter)")
    parser.add_argument("--seed", type=int, default=42, help="shuffle seed (stable reproducible)")
    parser.add_argument("--sheet_size", type=int, default=25, help="max rows per sheet")
    parser.add_argument("--prefer_audio", choices=["uk", "us"], default="uk")
    parser.add_argument("--timer", type=int, default=5000, help="timer ms (int)")
    parser.add_argument("--combo", type=int, default=2, help="repeat count (int)")
    parser.add_argument("--require_audio", action="store_true",
                        help="only keep entries that already have audio_primary (or fallback uk/us)")
    args = parser.parse_args()

    normalize_word = try_import_normalize() or fallback_normalize_word

    global_obj = load_json(args.global_cache) if args.source in ("global", "both") else {}
    uploaded_obj = load_json(args.uploaded_cache) if args.source in ("uploaded", "both") else {}

    raw_entries: List[Dict[str, Any]] = []
    if args.source in ("global", "both"):
        raw_entries.extend(iter_entries(global_obj))
    if args.source in ("uploaded", "both"):
        raw_entries.extend(iter_entries(uploaded_obj))

    # 补 word_norm（有些 entry 可能没有写入）
    for e in raw_entries:
        if not str(e.get("word_norm", "") or "").strip():
            wo = str(e.get("word_original", "") or "").strip()
            if wo:
                e["word_norm"] = normalize_word(wo)

    # 同词去重：取最完整的 entry
    picked = choose_best_entry_per_word(raw_entries)

    # 乱序
    rnd = random.Random(args.seed)
    rnd.shuffle(picked)

    # limit
    if args.limit and args.limit > 0:
        picked = picked[: args.limit]

    rows = build_rows(
        entries=picked,
        prefer_audio=args.prefer_audio,
        timer_ms=args.timer,
        combo_count=args.combo,
        require_audio=args.require_audio,
    )

    # 注意：build_rows 内可能因为 require_audio 过滤导致行数 < limit
    write_workbook(
        rows=rows,
        output_xlsx=args.output,
        sheet_size=args.sheet_size,
    )

    print(f"[DONE] source={args.source}, picked_words={len(picked)}, output_rows={len(rows)}, output={args.output}")


if __name__ == "__main__":
    main()
