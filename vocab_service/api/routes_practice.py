# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import os
import random
from pathlib import Path

from fastapi import APIRouter, HTTPException, Request

from api.schemas import (
    PracticeMasterRequest,
    PracticeMasterResponse,
    PracticeDocxRequest,
    PracticeDocxResponse,
)

from generators.excel_generator import (
    load_entries_from_vocab_excel,
    compute_available_practice_types,
    export_shuffle_e2c_excel_from_present_keys,
    ensure_audio_primary,
)

from datetime import datetime
from typing import Dict, List, Tuple

from openpyxl import load_workbook

# 复用你现有的 docx 生成核心函数（practice_excel_to_docx.py 里的内部函数）
from generators.practice_excel_to_docx import _build_one_docx_from_workbook, _get_header_keys

# ✅ 镜像 export：router 不带 prefix（由 app.include_router(..., prefix="/v1/practice") 决定）
router = APIRouter(tags=["practice"])


def _project_root() -> Path:
    # api/routes_practice.py -> api -> vocab_service(root)
    return Path(__file__).resolve().parents[1]


@router.post("/master", response_model=PracticeMasterResponse)
def build_practice_master(req: PracticeMasterRequest, request: Request):
    """
    生成 practice master xlsx（乱序单词英译中 master）
    镜像 export 思路：
    - 输入：vocab_note.xlsx（不回查 global_cache / run_cache）
    - 输出：storage/out/{run_id}/practice_master__{run_id}__{base_sheet_name}.xlsx
    - 同目录写 meta.json（包含 run_id + present_keys + available_practice_types）
    """
    rid = request.headers.get("x-request-id") or "N/A"

    ROOT = _project_root()

    # ===============================
    # 1) 校验 run_id / vocab_excel
    # ===============================
    run_id = (req.run_id or "").strip()
    if not run_id:
        raise HTTPException(status_code=400, detail="run_id is required")

    vocab_excel = (req.vocab_excel or "").strip()
    if not vocab_excel:
        raise HTTPException(status_code=400, detail="vocab_excel is required (must be exported vocab_note.xlsx)")

    vocab_excel_path = Path(vocab_excel).expanduser()
    if not vocab_excel_path.exists():
        raise HTTPException(status_code=400, detail=f"vocab_excel not found: {vocab_excel_path}")

    # ===============================
    # 2) 输出路径：统一归档到 storage/out/{run_id}/
    # ===============================
    out_dir = ROOT / "storage" / "out" / run_id
    out_dir.mkdir(parents=True, exist_ok=True)

    out_xlsx = (req.output_xlsx or "").strip()
    if not out_xlsx:
        out_xlsx = str((out_dir / f"practice_master__{run_id}__{req.base_sheet_name}.xlsx").resolve())

    meta_json = (req.meta_json or "").strip()
    if not meta_json:
        meta_json = str(Path(out_xlsx).with_suffix(".meta.json").resolve())

    # 确保父目录存在（如果用户传了别的路径）
    Path(out_xlsx).parent.mkdir(parents=True, exist_ok=True)
    Path(meta_json).parent.mkdir(parents=True, exist_ok=True)

    try:
        # ===============================
        # 3) 读取 vocab_note.xlsx -> entries + present_keys
        # ===============================
        pack = load_entries_from_vocab_excel(str(vocab_excel_path), sheet_name=req.sheet_name)
        entries = pack.get("entries", [])
        present_keys = pack.get("present_keys", [])

        if not entries:
            raise HTTPException(status_code=400, detail="vocab_excel entries is empty")

        # ===============================
        # 4) 决定可用练习类型 + 写 meta.json（镜像 export 的 meta 结构）
        # ===============================
        avail = compute_available_practice_types(present_keys)

        meta = {
            "run_id": run_id,
            "sheet_name": req.sheet_name,
            "source_vocab_excel": str(vocab_excel_path),
            "present_keys": list(present_keys),
            "available_practice_types": list(avail),
            "output_xlsx": out_xlsx,
            "base_sheet_name": req.base_sheet_name,
            "max_rows_per_sheet": req.max_rows_per_sheet,
            "seed": req.seed,
        }

        Path(meta_json).write_text(
            json.dumps(meta, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )

        # ===============================
        # 5) shuffle（seed=0 表示不固定）
        # ===============================
        entries2 = list(entries)
        if req.seed and req.seed != 0:
            rnd = random.Random(req.seed)
            rnd.shuffle(entries2)
        else:
            random.shuffle(entries2)

        # ===============================
        # 6) 确保 audio_primary
        # ===============================
        for e in entries2:
            if isinstance(e, dict):
                ensure_audio_primary(e)

        # ===============================
        # 7) 导出 master xlsx
        # ===============================
        export_shuffle_e2c_excel_from_present_keys(
            shuffled_entries=entries2,
            present_keys=present_keys,
            output_path=out_xlsx,
            base_sheet_name=req.base_sheet_name,
            max_rows_per_sheet=req.max_rows_per_sheet,
        )

        # ===============================
        # 8) 返回（字段风格尽量对齐 export：run_id + 产物路径 + count）
        # ===============================
        return PracticeMasterResponse(
            run_id=run_id,
            vocab_excel=str(vocab_excel_path),
            output_xlsx=out_xlsx,
            meta_json=meta_json,
            rows=len(entries2),
            present_keys=list(present_keys),
            available_practice_types=list(avail),
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"[{rid}] practice master failed: {e}")


def _project_root() -> Path:
    return Path(__file__).resolve().parents[1]  # vocab_service


def _stamp(ts: str, run_id: str) -> str:
    ts2 = (ts or "").strip()
    rid = (run_id or "").strip()
    if ts2:
        return f"{ts2}_{rid}" if rid else ts2
    ts2 = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{ts2}_{rid}" if rid else ts2

def _specs(suffix: str) -> Dict[str, Dict]:
    """
    与 generators/practice_excel_to_docx.py 的四类练习保持一致
    """
    DOUBLE_WORD_COL_WIDTHS_CM = [1.0, 3.8, 3.8, 1.0, 3.8, 3.8]  # A 方案：竖版不溢出

    return {
        "word_e2c": {
            "title": "单词 英译中（学生版）",
            "filename": f"练习_单词_英译中_{suffix}.docx",
            "layout": "double",
            "col_defs": [
                {"key": "no", "blank": False},
                {"key": "word_original", "blank": False},
                {"key": "pos_cn", "blank": True},
            ],
            "col_widths_cm": DOUBLE_WORD_COL_WIDTHS_CM,
            "body_row_height_cm": 1.0,   # ✅ 1.0cm
        },
        "word_c2e": {
            "title": "单词 中译英（学生版）",
            "filename": f"练习_单词_中译英_{suffix}.docx",
            "layout": "double",
            "col_defs": [
                {"key": "no", "blank": False},
                {"key": "pos_cn", "blank": False},
                {"key": "word_original", "blank": True},
            ],
            "col_widths_cm": DOUBLE_WORD_COL_WIDTHS_CM,
            "body_row_height_cm": 1.0,   # ✅ 1.0cm
        },
        "sent_e2c": {
            "title": "例句 英译中（学生版）",
            "filename": f"练习_例句_英译中_{suffix}.docx",
            "layout": "single",
            "col_defs": [
                {"key": "no", "blank": False},
                {"key": "example", "blank": False},
                {"key": "example_cn", "blank": True},
            ],
            "col_widths_cm": [1.2, 9.0, 9.0],
            "body_row_height_cm": 2.0,
        },
        "sent_c2e": {
            "title": "例句 中译英（学生版）",
            "filename": f"练习_例句_中译英_{suffix}.docx",
            "layout": "single",
            "col_defs": [
                {"key": "no", "blank": False},
                {"key": "example_cn", "blank": False},
                {"key": "example", "blank": True},
            ],
            "col_widths_cm": [1.2, 9.0, 9.0],
            "body_row_height_cm": 2.0,
        },
    }


@router.post("/docx", response_model=PracticeDocxResponse)
def build_practice_docx(req: PracticeDocxRequest, request: Request):
    """
    镜像 export 的风格：
    - 输入：master_xlsx（表头为 key）
    - 输出：storage/out/{run_id}/practice_docx/*.docx
    - 缺字段就跳过（与你脚本一致）
    """
    rid = request.headers.get("x-request-id") or "N/A"

    run_id = (req.run_id or "").strip()
    if not run_id:
        raise HTTPException(status_code=400, detail="run_id is required")

    master_xlsx = (req.master_xlsx or "").strip()
    if not master_xlsx:
        raise HTTPException(status_code=400, detail="master_xlsx is required")

    master_path = Path(master_xlsx).expanduser()
    if not master_path.exists():
        raise HTTPException(status_code=400, detail=f"master_xlsx not found: {master_path}")

    ROOT = _project_root()

    # 输出目录：默认 storage/out/{run_id}/practice_docx/
    out_dir = (req.output_dir or "").strip()
    if not out_dir:
        out_dir = str((ROOT / "storage" / "out" / run_id / "practice_docx").resolve())

    Path(out_dir).mkdir(parents=True, exist_ok=True)

    suffix = _stamp(req.timestamp, run_id)
    all_specs = _specs(suffix)

    # 选择 types：空 -> 全部
    types = req.types or []
    if not types:
        types = ["word_e2c", "word_c2e", "sent_e2c", "sent_c2e"]

    # 校验 types 合法
    for t in types:
        if t not in all_specs:
            raise HTTPException(status_code=400, detail=f"unknown type: {t}")

    try:
        wb = load_workbook(str(master_path), data_only=True)
        first_ws = wb[wb.sheetnames[0]]
        header_keys = set(_get_header_keys(first_ws))

        generated: List[str] = []
        skipped: List[str] = []

        for t in types:
            sp = all_specs[t]
            required = [c["key"] for c in sp["col_defs"] if not c.get("blank", False)]
            missing = [k for k in required if k not in header_keys]

            if missing:
                skipped.append(t)
                continue

            out_path = str((Path(out_dir) / sp["filename"]).resolve())
            _build_one_docx_from_workbook(
                wb=wb,
                output_path=out_path,
                doc_title=sp["title"],
                col_defs=sp["col_defs"],
                col_widths_cm=sp["col_widths_cm"],
                body_row_height_cm=sp["body_row_height_cm"],
                layout=sp.get("layout", "single"),
            )
            generated.append(out_path)

        return PracticeDocxResponse(
            run_id=run_id,
            master_xlsx=str(master_path),
            output_dir=str(Path(out_dir).resolve()),
            generated=generated,
            skipped=skipped,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"[{rid}] practice docx failed: {e}")
