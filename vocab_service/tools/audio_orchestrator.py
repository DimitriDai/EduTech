# tools/audio_orchestrator.py
# -*- coding: utf-8 -*-

import argparse
import os
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def _safe_filename(s: str) -> str:
    s = (s or "").strip()
    # Windows illegal chars: \ / : * ? " < > |
    s = re.sub(r'[\\/:*?"<>|]+', "_", s)
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "sheet"


def _run(cmd: list[str]) -> int:
    print("[RUN]", " ".join(f'"{x}"' if " " in x else x for x in cmd))
    p = subprocess.run(cmd, shell=False)
    return p.returncode


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input_excel", required=True, help="xlsx with key headers, may contain multiple sheets")
    parser.add_argument("--output_dir", required=True, help="directory to write full + parts mp3")
    parser.add_argument(
        "--compose_script",
        default=str(Path(__file__).resolve().parent / "practice_audio_compose.py"),
        help="path to practice_audio_compose.py",
    )

    # pass-through params (keep consistent with practice_audio_compose.py)
    parser.add_argument("--audio_root", default="storage/audio_cache")
    parser.add_argument("--ffmpeg_bin", default="ffmpeg")
    parser.add_argument("--ffprobe_bin", default="ffprobe")
    parser.add_argument("--silence_cache_dir", default="storage/audio_cache/_silence")
    parser.add_argument("--no_manifest", action="store_true", help="disable manifest output")

    # naming
    parser.add_argument("--prefix", default="", help="filename prefix (optional)")
    parser.add_argument("--make_parts", action="store_true", help="also generate per-sheet part mp3 files")
    args = parser.parse_args()

    input_excel = Path(args.input_excel).resolve()
    if not input_excel.exists():
        raise SystemExit(f"[FATAL] input_excel not found: {input_excel}")

    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    compose_script = Path(args.compose_script).resolve()
    if not compose_script.exists():
        raise SystemExit(f"[FATAL] compose_script not found: {compose_script}")

    wb = load_workbook(str(input_excel), data_only=True, read_only=True)
    sheetnames = wb.sheetnames
    if not sheetnames:
        raise SystemExit("[FATAL] No sheets found in workbook")

    stem = input_excel.stem
    prefix = args.prefix.strip()
    if prefix:
        base = f"{prefix}_{stem}"
    else:
        base = stem

    py = sys.executable

    # 1) FULL (all sheets)
    full_mp3 = output_dir / f"{base}_full.mp3"
    cmd_full = [
        py, str(compose_script),
        "--input_excel", str(input_excel),
        "--audio_root", args.audio_root,
        "--output_mp3", str(full_mp3),
        "--ffmpeg_bin", args.ffmpeg_bin,
        "--ffprobe_bin", args.ffprobe_bin,
        "--silence_cache_dir", args.silence_cache_dir,
    ]
    if args.no_manifest:
        cmd_full.append("--no_manifest")

    rc = _run(cmd_full)
    if rc != 0:
        raise SystemExit(f"[FATAL] FULL generation failed with code={rc}")

    print(f"[OK] FULL -> {full_mp3}")

    # 2) PARTS (each sheet)
    if args.make_parts:
        for idx, sh in enumerate(sheetnames, start=1):
            safe = _safe_filename(sh)
            part_mp3 = output_dir / f"{base}_part_{idx:02d}_{safe}.mp3"

            cmd_part = [
                py, str(compose_script),
                "--input_excel", str(input_excel),
                "--only_sheets", sh,  # practice_audio_compose expects comma-separated; single is ok
                "--audio_root", args.audio_root,
                "--output_mp3", str(part_mp3),
                "--ffmpeg_bin", args.ffmpeg_bin,
                "--ffprobe_bin", args.ffprobe_bin,
                "--silence_cache_dir", args.silence_cache_dir,
            ]
            if args.no_manifest:
                cmd_part.append("--no_manifest")

            rc2 = _run(cmd_part)
            if rc2 != 0:
                raise SystemExit(f"[FATAL] PART generation failed (sheet={sh}) code={rc2}")

            print(f"[OK] PART  ({sh}) -> {part_mp3}")

    print("[DONE] audio_orchestrator finished.")


if __name__ == "__main__":
    main()