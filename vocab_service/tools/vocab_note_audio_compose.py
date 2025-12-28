# -*- coding: utf-8 -*-
"""
tools/vocab_note_audio_compose.py

用途：
- 基于 excel_generator.py 产出的 vocab_note.xlsx（正序）生成整段音频
- 不修改原表格
- 默认 combo=2, timer=4000（仅当表格没有提供 timer/combo 时）

输出：
- {output_dir}/{prefix}_vocab_note_full.mp3
- 可选：manifest.json（取决于 practice_audio_compose.py 的参数）
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def _repo_root() -> Path:
    # tools/xxx.py -> vocab_service
    return Path(__file__).resolve().parents[1]


def _find_header(ws) -> list[str]:
    rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not rows:
        return []
    return [str(x).strip() if x is not None else "" for x in rows[0]]


def _has_any_data(ws, col_idx: int, max_check: int = 50) -> bool:
    # 从第2行起检查是否有非空
    checked = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if col_idx < len(r):
            v = r[col_idx]
            if v is not None and str(v).strip() != "":
                return True
        checked += 1
        if checked >= max_check:
            break
    return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--vocab_excel", required=True, help="vocab_note.xlsx 路径（正序）")
    ap.add_argument("--sheet", default="vocab", help="sheet 名（默认 vocab）")
    ap.add_argument("--output_dir", default="", help="输出目录（默认 storage/out/_adhoc/audio）")
    ap.add_argument("--prefix", default="", help="输出文件名前缀（默认使用 vocab_excel 文件名 stem）")
    ap.add_argument("--audio_root", default="", help="音频根目录（默认 storage/audio_cache）")
    ap.add_argument("--ffmpeg_bin", default="", help="ffmpeg.exe 全路径（可空）")
    ap.add_argument("--ffprobe_bin", default="", help="ffprobe.exe 全路径（可空）")

    # ✅ 你要的默认值
    ap.add_argument("--default_timer_ms", type=int, default=4000)
    ap.add_argument("--default_combo", type=int, default=2)

    # 是否输出 manifest
    ap.add_argument("--no_manifest", action="store_true")

    args = ap.parse_args()

    ROOT = _repo_root()
    vocab_path = Path(args.vocab_excel).expanduser().resolve()
    if not vocab_path.exists():
        raise SystemExit(f"[FATAL] vocab_excel not found: {vocab_path}")

    wb = load_workbook(str(vocab_path), data_only=True)
    sheet = args.sheet if args.sheet in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    header = _find_header(ws)
    header_l = [h.lower() for h in header]

    # 最低要求：word_original & audio_primary
    def _col(name: str) -> int:
        return header_l.index(name)

    if "word_original" not in header_l:
        raise SystemExit("[FATAL] vocab_note.xlsx missing column: word_original")
    if "audio_primary" not in header_l:
        # vocab_note 未输出 audio_primary 时，直接失败（因为不改原表格）
        raise SystemExit("[FATAL] vocab_note.xlsx missing column: audio_primary (export 时需包含音频列)")

    # 检查是否有数据
    if not _has_any_data(ws, _col("word_original")):
        raise SystemExit("[FATAL] vocab_note.xlsx has no word data")

    # 输出目录默认：storage/out/_adhoc/audio（如果你希望必须 run_id，也可以在 API 层拼好传进来）
    out_dir = Path(args.output_dir).expanduser() if args.output_dir else (ROOT / "storage" / "out" / "_adhoc" / "audio")
    out_dir.mkdir(parents=True, exist_ok=True)

    prefix = args.prefix.strip() or vocab_path.stem

    compose_py = (ROOT / "tools" / "practice_audio_compose.py").resolve()
    if not compose_py.exists():
        raise SystemExit(f"[FATAL] practice_audio_compose.py not found: {compose_py}")

    # 调用 compose 脚本：它会读取 timer/combo，如果缺失则使用默认值
    # 这里我们通过参数把默认值“传给 compose”，前提是 compose 支持；如果你现有 compose 没参数，就改为：
    # - 在调用前生成一个临时 xlsx（只用于音频），写 timer/combo 两列（仍不改原表格）
    cmd = [
        sys.executable,
        str(compose_py),
        "--input_excel", str(vocab_path),
        "--output_dir", str(out_dir),
        "--sheet", sheet,
        "--prefix", f"{prefix}_vocab_note",
        "--default_timer_ms", str(args.default_timer_ms),
        "--default_combo", str(args.default_combo),
    ]

    if args.audio_root:
        cmd += ["--audio_root", args.audio_root]
    if args.ffmpeg_bin:
        cmd += ["--ffmpeg_bin", args.ffmpeg_bin]
    if args.ffprobe_bin:
        cmd += ["--ffprobe_bin", args.ffprobe_bin]
    if args.no_manifest:
        cmd += ["--no_manifest"]

    # 执行
    p = subprocess.run(cmd, cwd=str(ROOT), text=True, capture_output=True)
    if p.returncode != 0:
        tail = (p.stderr or p.stdout or "").strip()[-1200:]
        raise SystemExit(f"[FATAL] compose failed: {tail}")

    print(f"[DONE] vocab_note audio composed in: {out_dir}")


if __name__ == "__main__":
    main()
